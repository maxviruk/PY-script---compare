import os
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === CONFIGURATION ===
watch_dir = os.path.join(os.getcwd(), "PY - Data - EOPWD")
log_dir = os.path.join(os.getcwd(), "PY - Logs")
os.makedirs(log_dir, exist_ok=True)

max_valid_date = pd.Timestamp("2262-04-11")

file_sap = "Table_SAP.xlsx"
# WD will be discovered automatically by prefix:
WD_FILE_PREFIX = "Table_WD"   # matches Table_WD.xlsx and Table_WD_*.xlsx

output_file = "SAP_Expanded.xlsx"      # neutral name
log_file = "processing_log_1.txt"
check_interval = 10  # seconds

# Columns we keep as-is from SAP; others become "-"
required_columns = [
    "Pers.No.", "Personnel Number", "EEGrp", "Employee Group", "S", "Employment Status",
    "CoCd", "Company Code", "PA", "Personnel Area", "ESgrp", "Employee Subgroup",
    "Start Date", "End Date", "Changed by", "Start", "End time",
    "A/AType", "Attendance or Absence Type"
]

def log(msg: str) -> None:
    """Append a timestamped line to the log file and print it."""
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(os.path.join(log_dir, log_file), "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {msg}\n")
    print(f"[{ts}] {msg}")

def get_unique_output_path(base_dir: str, base_filename: str) -> str:
    """
    Return a non-colliding output path inside base_dir.
    If <name>.xlsx exists, try <name>_<ddmmyyyy>.xlsx, then <name>_<ddmmyyyy>-N.xlsx.
    """
    base_path = os.path.join(base_dir, base_filename)
    if not os.path.exists(base_path):
        return base_path

    name, ext = os.path.splitext(base_filename)
    date_suffix = datetime.now().strftime("%d%m%Y")
    candidate = os.path.join(base_dir, f"{name}_{date_suffix}{ext}")
    if not os.path.exists(candidate):
        return candidate

    counter = 1
    while True:
        candidate = os.path.join(base_dir, f"{name}_{date_suffix}-{counter}{ext}")
        if not os.path.exists(candidate):
            return candidate
        counter += 1

def add_formula_columns(xlsx_path: str) -> None:
    """Ensure 'PY' column exists and mark rows as 'Compared'."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    if "PY" not in headers:
        col_idx = ws.max_column + 1
        ws.cell(row=1, column=col_idx).value = "PY"
        col_letter = get_column_letter(col_idx)
        for r in range(2, ws.max_row + 1):
            ws[f"{col_letter}{r}"] = "Compared"

    wb.save(xlsx_path)
    wb.close()

def reorder_columns(xlsx_path: str) -> None:
    """
    Move service columns (if present) to the far right, preserving data.
    Only moves columns that actually exist in the sheet.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    cols_to_move = ["AbsenceDate_SAP", "Key_SAP", "PY"]
    indices_to_move = [headers.index(c) + 1 for c in cols_to_move if c in headers]

    max_col = ws.max_column
    max_row = ws.max_row
    new_col_start = max_col + 1

    # Copy columns to the end
    for i, src_idx in enumerate(indices_to_move):
        dst_idx = new_col_start + i
        ws.cell(row=1, column=dst_idx).value = ws.cell(row=1, column=src_idx).value
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=dst_idx).value = ws.cell(row=r, column=src_idx).value

    # Delete originals (from right to left)
    for src_idx in sorted(indices_to_move, reverse=True):
        ws.delete_cols(src_idx)

    wb.save(xlsx_path)
    wb.close()

def add_formula_and_remove_duplicates(xlsx_path: str, temp_title: str = "#") -> None:
    """
    Add a temporary formula column for duplicate detection, remove dup rows, then drop the temp column.
    Formula is kept as-is from your previous version.
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    if temp_title not in headers:
        col_idx = ws.max_column + 1
        ws.cell(row=1, column=col_idx).value = temp_title
    else:
        col_idx = headers.index(temp_title) + 1

    col_letter = get_column_letter(col_idx)
    base_formula = '=IFERROR(IF(AK2="SAP Payroll systems";CONCAT(A2;V2;M2);"-");"-")'

    for r in range(2, ws.max_row + 1):
        ws[f"{col_letter}{r}"] = base_formula.replace("2", str(r))

    # Read back values to detect duplicates quickly
    seen = set()
    rows_to_delete = []
    for r in range(2, ws.max_row + 1):
        val = ws[f"{col_letter}{r}"].value
        if val in seen:
            rows_to_delete.append(r)
        else:
            seen.add(val)

    for r in reversed(rows_to_delete):
        ws.delete_rows(r)

    ws.delete_cols(col_idx)

    wb.save(xlsx_path)
    wb.close()
    log(f"✅ Removed {len(rows_to_delete)} duplicates using temp column '{temp_title}'.")

def find_latest_wd_file(dir_path: str, prefix: str = "Table_WD") -> str | None:
    """
    Return absolute path to the newest WD file in dir_path matching:
      - starts with 'Table_WD'
      - ends with '.xlsx'
    Ignores temporary files like '~$...xlsx'.
    """
    latest = None
    latest_mtime = -1
    for name in os.listdir(dir_path):
        if name.startswith("~$"):
            continue
        if name.startswith(prefix) and name.lower().endswith(".xlsx"):
            p = os.path.join(dir_path, name)
            if os.path.isfile(p):
                m = os.path.getmtime(p)
                if m > latest_mtime:
                    latest = p
                    latest_mtime = m
    return latest

def process_files(wd_path: str) -> None:
    """Main worker: load SAP and the newest WD file, expand SAP by day (ALL absence types), save, post-process."""
    try:
        out_path = get_unique_output_path(watch_dir, output_file)
        log(f"📁 Saving result to: {os.path.basename(out_path)}")

        sap_path = os.path.join(watch_dir, file_sap)

        log("📂 Loading input files")
        sap_df = pd.read_excel(sap_path)
        wd_df  = pd.read_excel(wd_path)
        log(f"ℹ️ Using WD file: {os.path.basename(wd_path)}")

        # Clean possible leftovers from previous runs
        for c in ["AbsenceDate_SAP", "Key_SAP"]:
            if c in sap_df.columns:
                sap_df.drop(columns=[c], inplace=True)

        # === DATE SANITIZATION (prevents OutOfBounds for 9999-12-31 etc.) ===
        sap_df["Start Date"] = pd.to_datetime(sap_df["Start Date"], errors="coerce")
        sap_df["End Date"]   = pd.to_datetime(sap_df["End Date"],   errors="coerce")
        sap_df = sap_df[
            sap_df["Start Date"].notna()
            & sap_df["End Date"].notna()
            & (sap_df["End Date"] <= max_valid_date)
        ]

        # Build WD Key (kept from your version; not used further but harmless)
        if {"Employee ID", "Time Off date"}.issubset(wd_df.columns):
            wd_df["Time Off date"] = pd.to_datetime(wd_df["Time Off date"], errors="coerce")
            wd_df["Key_WD"] = wd_df["Employee ID"].astype(str) + "_" + wd_df["Time Off date"].dt.strftime("%Y%m%d")

        all_columns = sap_df.columns.tolist()
        rows = []

        def build_row(src_row, overrides: dict):
            # Keep specific SAP columns; others as "-"
            full_row = {col: src_row[col] if col in required_columns else "-" for col in all_columns}
            full_row.update(overrides)
            return full_row

        # Expand EVERY SAP row by day (for all absence types)
        for _, row in sap_df.iterrows():
            start = row["Start Date"]
            end = row["End Date"]

            if start == end:
                rows.append(build_row(row, {
                    "Start Date": start,
                    "End Date": end,
                    "AbsenceDate_SAP": start,
                    "Key_SAP": f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}",
                    "PY": None
                }))
            else:
                for d in pd.date_range(start, end, freq="D"):
                    rows.append(build_row(row, {
                        "Start Date": d,
                        "End Date": d,
                        "AbsenceDate_SAP": d,
                        "Key_SAP": f"{row['Personnel Number']}_{d.strftime('%Y%m%d')}",
                        "PY": None
                    }))

        df = pd.DataFrame(rows)

        # Mark PY where it was None
        if "PY" in df.columns:
            df["PY"] = df["PY"].apply(lambda x: "Python script" if pd.isna(x) else x)

        # Unique by Key_SAP
        if "Key_SAP" in df.columns:
            df = df.drop_duplicates(subset=["Key_SAP"], keep="first")

        # Normalize time fields
        for c in ["Start", "End time"]:
            if c in df.columns:
                df[c] = df[c].astype(str).replace({":  :": "-"})

        # Put service columns to the end (if present)
        cols = df.columns.tolist()
        for c in ["AbsenceDate_SAP", "Key_SAP", "PY"]:
            if c in cols:
                cols.remove(c)
        cols.extend([c for c in ["AbsenceDate_SAP", "Key_SAP", "PY"] if c in df.columns])
        df = df[cols]

        # Drop service columns before saving (as in your original flow)
        for c in ["AbsenceDate_SAP", "Key_SAP"]:
            if c in df.columns:
                df.drop(columns=[c], inplace=True)

        log("📊 Saving results to Excel")
        df.to_excel(out_path, index=False)

        # Post-processing in the saved workbook
        add_formula_columns(out_path)
        reorder_columns(out_path)                 # will only move what exists (likely just 'PY')
        add_formula_and_remove_duplicates(out_path, "#")

        log("✅ Processing completed successfully.")
    except Exception as e:
        log(f"❌ ERROR during processing: {e}")

def wait_for_files() -> None:
    """Block until SAP exists and any Table_WD*.xlsx exists; then process the newest WD."""
    log("🚀 Script started. Waiting for files")
    log("🔍 Watching for input files")
    while True:
        files = {f.lower() for f in os.listdir(watch_dir)}
        sap_ok = "table_sap.xlsx" in files
        wd_path = find_latest_wd_file(watch_dir, WD_FILE_PREFIX)
        if sap_ok and wd_path:
            log(f"📂 Detected SAP and WD files. Latest WD: {os.path.basename(wd_path)}. Starting processing.")
            process_files(wd_path)
            break
        time.sleep(check_interval)

if __name__ == "__main__":
    wait_for_files()
