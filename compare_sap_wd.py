import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === Step 1: Load data ===
sap_df = pd.read_excel("Table_SAP.xlsx")
wd_df = pd.read_excel("Table_WD.xlsx")

# === Step 2: Filter AS01 only ===
sap_df = sap_df[sap_df["A/AType"] == "AS01"].copy()

# === Step 3: Set required columns ===
required_columns = [
    "Pers.No.", "Personnel Number", "EEGrp", "Employee Group", "S", "Employment Status",
    "CoCd", "Company Code", "PA", "Personnel Area", "ESgrp", "Employee Subgroup",
    "Start Date", "End Date", "Start", "End time", "A/AType", "Attendance or Absence Type"
]
all_columns = sap_df.columns.tolist()

# === Step 4: Expand SAP with logic ===
rows = []

for _, row in sap_df.iterrows():
    start = row["Start Date"]
    end = row["End Date"]

    if pd.isnull(start) or pd.isnull(end):
        continue

    if start == end:
        # 1-day absence, single row
        one_day = {col: row[col] if col in required_columns else None for col in all_columns}
        one_day["Start Date"] = start
        one_day["End Date"] = end
        one_day["AbsenceDate_SAP"] = start
        one_day["Key_SAP"] = f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}"
        one_day["Status"] = None
        rows.append(one_day)
    else:
        # Original row
        original = {col: row[col] if col in required_columns else None for col in all_columns}
        original["AbsenceDate_SAP"] = pd.NaT
        original["Key_SAP"] = f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}"
        original["Status"] = "ORIGINAL"
        rows.append(original)

        # Split rows
        for d in pd.date_range(start, end):
            r = {col: row[col] if col in required_columns else None for col in all_columns}
            r["Start Date"] = d
            r["End Date"] = d
            r["AbsenceDate_SAP"] = d
            r["Key_SAP"] = f"{row['Personnel Number']}_{d.strftime('%Y%m%d')}"
            r["Status"] = None
            rows.append(r)

df = pd.DataFrame(rows)

# === Step 5: WD key and comparison ===
wd_df["Key_WD"] = wd_df["Employee ID"].astype(str) + "_" + wd_df["Time Off date"].dt.strftime("%Y%m%d")
df["Status"] = df["Key_SAP"].isin(wd_df["Key_WD"]).map({True: "OK", False: "Missing in WD"}).where(df["Status"] != "ORIGINAL", "ORIGINAL")

# === Step 6: Remove duplicates ===
df = df.drop_duplicates(subset=["Key_SAP", "Status"], keep="first")
df = df.drop_duplicates(subset=["Key_SAP"], keep="first")

# === Step 7: Save and highlight ===
output_file = "SAP_vs_WD_Final.xlsx"
df.to_excel(output_file, index=False)

# Highlight ORIGINAL rows in yellow
wb = load_workbook(output_file)
ws = wb.active
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
status_col = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}.get("Status", None)

if status_col:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[status_col - 1].value == "ORIGINAL":
            for cell in row:
                cell.fill = fill

wb.save(output_file)
wb.close()
