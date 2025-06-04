import os
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === CONFIG ===
watch_dir = os.path.dirname(os.path.abspath(__file__))
file_sap = "Table_SAP.xlsx"
file_wd = "Table_WD.xlsx"
output_file = "SAP_vs_WD_Final.xlsx"
log_file = "processing_log.txt"
check_interval = 10  # seconds

required_columns = [
    "Pers.No.", "Personnel Number", "EEGrp", "Employee Group", "S", "Employment Status",
    "CoCd", "Company Code", "PA", "Personnel Area", "ESgrp", "Employee Subgroup",
    "Start Date", "End Date", "Start", "End time", "A/AType", "Attendance or Absence Type"
]

def log(msg):
    with open(os.path.join(watch_dir, log_file), "a", encoding="utf-8") as logf:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logf.write(f"[{timestamp}] {msg}\n")
    print(f"[{timestamp}] {msg}")

def process_files():
    try:
        sap_path = os.path.join(watch_dir, file_sap)
        wd_path = os.path.join(watch_dir, file_wd)
        out_path = os.path.join(watch_dir, output_file)

        sap_df = pd.read_excel(sap_path)
        wd_df = pd.read_excel(wd_path)

        sap_df = sap_df[sap_df["A/AType"] == "AS01"].copy()
        all_columns = sap_df.columns.tolist()
        wd_df["Key_WD"] = wd_df["Employee ID"].astype(str) + "_" + wd_df["Time Off date"].dt.strftime("%Y%m%d")

        rows = []

        for _, row in sap_df.iterrows():
            start = row["Start Date"]
            end = row["End Date"]
            if pd.isnull(start) or pd.isnull(end):
                continue

            if start == end:
                one_day = {col: row[col] if col in required_columns else None for col in all_columns}
                one_day["Start Date"] = start
                one_day["End Date"] = end
                one_day["AbsenceDate_SAP"] = start
                one_day["Key_SAP"] = f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}"
                one_day["Status"] = None
                rows.append(one_day)
            else:
                original = {col: row[col] if col in required_columns else None for col in all_columns}
                original["AbsenceDate_SAP"] = pd.NaT
                original["Key_SAP"] = f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}"
                original["Status"] = "ORIGINAL"
                rows.append(original)

                for d in pd.date_range(start, end):
                    r = {col: row[col] if col in required_columns else None for col in all_columns}
                    r["Start Date"] = d
                    r["End Date"] = d
                    r["AbsenceDate_SAP"] = d
                    r["Key_SAP"] = f"{row['Personnel Number']}_{d.strftime('%Y%m%d')}"
                    r["Status"] = None
                    rows.append(r)

        df = pd.DataFrame(rows)
        df["Status"] = df["Key_SAP"].isin(wd_df["Key_WD"]).map({True: "OK", False: "Missing in WD"}).where(df["Status"] != "ORIGINAL", "ORIGINAL")
        df = df.drop_duplicates(subset=["Key_SAP", "Status"], keep="first")
        df = df.drop_duplicates(subset=["Key_SAP"], keep="first")

        df.to_excel(out_path, index=False)

        # Highlight
        wb = load_workbook(out_path)
        ws = wb.active
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        status_col = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}.get("Status", None)

        if status_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[status_col - 1].value == "ORIGINAL":
                    for cell in row:
                        cell.fill = fill

        wb.save(out_path)
        wb.close()

        log("✅ Processing completed successfully.")
    except Exception as e:
        log(f"❌ ERROR during processing: {e}")

def wait_for_files():
    log("🔍 Watching for input files...")
    while True:
        files = os.listdir(watch_dir)
        if file_sap in files and file_wd in files:
            log("📂 Detected both files. Starting processing.")
            process_files()
            break
        time.sleep(check_interval)

if __name__ == "__main__":
    wait_for_files()
