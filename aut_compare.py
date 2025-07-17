import os
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter



# === CONFIGURATION and required fields keep/populate ===
watch_dir = os.path.join(os.getcwd(), "PY - Data - EOPWD")
log_dir = os.path.join(os.getcwd(), "PY - Logs")
max_valid_date = pd.Timestamp("2262-04-11")
file_sap = "Table_SAP.xlsx"                              # SAP input file
file_wd = "Table_WD.xlsx"                                # Workday input file
output_file = "AS01,AX04,AS03,AH01 - SAP_vs_WD.xlsx"     # Output Excel file
log_file = "processing_log_1.txt"                        # Log file
check_interval = 10                                      # Seconds to wait before checking for files again
required_columns = [
    "Pers.No.", "Personnel Number", "EEGrp", "Employee Group", "S", "Employment Status",
    "CoCd", "Company Code", "PA", "Personnel Area", "ESgrp", "Employee Subgroup", "Start Date",
    "End Date", "Changed by", "Start", "End time", "A/AType", "Attendance or Absence Type"
]



# === F1 - Check if output file exists ===
def get_unique_output_path(watch_dir, base_filename):
    base_path = os.path.join(watch_dir, base_filename)
    if not os.path.exists(base_path):
        return base_path
    
    name, ext = os.path.splitext(base_filename)
    date_suffix = datetime.now().strftime("%d%m%Y")
    dated_name = f"{name}_{date_suffix}{ext}"
    dated_path = os.path.join(watch_dir, dated_name)

    if not os.path.exists(dated_path):
        return dated_path
    
    counter = 1
    while True:
        candidate_name = f"{name}_{date_suffix}-{counter}{ext}"
        candidate_path = os.path.join(watch_dir, candidate_name)
        if not os.path.exists(candidate_path):
            return candidate_path
        counter += 1



# === F2 - Logging utility ===
def log(msg):
    with open(os.path.join(log_dir, log_file), "a", encoding="utf-8") as logf:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logf.write(f"[{timestamp}] {msg}\n")
    print(f"[{timestamp}] {msg}")



# === F3 - Add formula columns to file ===
def add_formula_columns(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    max_row = ws.max_row
    last_col_idx = ws.max_column

    new_columns = [
        ("Country ID",            '=LEFT(G{row},2)'),
        ("Country3",              '=VLOOKUP(G{row},Integration!$M:$O,3,0)'),
        ("SAP / non-SAP",         '=IF(AJ{row}="-","-",VLOOKUP(AJ{row},Integration!$B:$I,8,0))'),
        ("#", '''=IF(AND(AK{row}="Non SAP Non Payroll systems",LEN(V{row})=0),"Non SAP Non Payroll systems",
    IF(AK{row}="Non SAP Non Payroll systems",CONCAT(A{row},V{row},M{row}),
    IF(AK{row}="SAP Payroll systems",CONCAT(A{row},V{row},M{row}),"-")))'''),
        ("##", '''=IF(AND(AK{row}="Non SAP Non Payroll systems",LEN(V{row})=0),"Non SAP Non Payroll systems",
    IF(AK{row}="Non SAP Non Payroll systems",CONCAT(A{row},V{row},M{row},N{row}),
    IF(AK{row}="SAP Payroll systems",CONCAT(A{row},V{row},M{row},N{row}),"-")))'''),
        ("CHECK", '''=IFERROR(
    IF(YEAR(M{row})=2024,AL{row},
    VLOOKUP(AL{row},WD!$BB:$BB,1,0)),"-")'''),
        ("OK/NOK", '''=IFERROR(
    IF(YEAR(M{row})=2024,"OK",
    IF(OR(AN{row}<>"-",AP{row}="Integration"),"OK","NOK")),"NOK")'''),
        ("CHECK - Changed by - L1", None),
        ("CHECK - Changed by - L2", None),
    ]

    for i, (col_name, formula_template) in enumerate(new_columns):
        col_idx = last_col_idx + 1 + i
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}1"] = col_name
        ws[f"{col_letter}1"].alignment = Alignment(horizontal='center')
        for row in range(2, max_row + 1):
            if formula_template:
                ws[f"{col_letter}{row}"] = formula_template.format(row=row)

    wb.save(file_path)
    wb.close()



# === F4 - reorder columns in Excel ===
def reorder_columns(file_path):
    wb = load_workbook(file_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    cols_to_move = ["AbsenceDate_SAP", "Key_SAP", "PY"]
    indices_to_move = [headers.index(col) + 1 for col in cols_to_move if col in headers]
    max_col = ws.max_column
    max_row = ws.max_row
    new_col_start = max_col + 1

    #=== F4.1 - Copy the contents of the columns to the end ===
    for i, col_idx in enumerate(indices_to_move):
        new_col = new_col_start + i
        ws.cell(row=1, column=new_col).value = ws.cell(row=1, column=col_idx).value
        for row in range(2, max_row + 1):
            ws.cell(row=row, column=new_col).value = ws.cell(row=row, column=col_idx).value

    #=== F4.1 - Delete the old columns from right to left to avoid messing up the index ===
    for col_idx in sorted(indices_to_move, reverse=True):
        ws.delete_cols(col_idx)

    wb.save(file_path)
    wb.close()



# === F5 - Main processing function ===
def process_files():
    try:
        out_path = get_unique_output_path(watch_dir, output_file)
        log(f"📁 Saving result to: {os.path.basename(out_path)}")

        sap_path = os.path.join(watch_dir, file_sap)
        wd_path = os.path.join(watch_dir, file_wd)

        log("📂 Loading input files...")
        sap_df = pd.read_excel(sap_path)
        wd_df = pd.read_excel(wd_path)


        # === F5.1 - Remove unnecessary columns if they exist ===
        for col_to_drop in ["AbsenceDate_SAP", "Key_SAP"]:
            if col_to_drop in sap_df.columns:
                sap_df.drop(columns=[col_to_drop], inplace=True)


        # Filter for absence type below:
        # "AS01", "AX04","AS03" - GE
        # "AH01" - LUX
        sap_df = sap_df[sap_df["A/AType"].isin(["AS01", "AX04", "AS03", "AH01"])].copy()
        all_columns = sap_df.columns.tolist()
        wd_df["Key_WD"] = wd_df["Employee ID"].astype(str) + "_" + wd_df["Time Off date"].dt.strftime("%Y%m%d")
        rows = []


        for _, row in sap_df.iterrows():
            start = row["Start Date"]
            end = row["End Date"]
            if pd.isnull(start) or pd.isnull(end):
                continue
            
            
            # === F5.1 - Fallback row generator with dashes for optional fields ===
            def build_row(overrides):
                full_row = {col: row[col] if col in required_columns else "-" for col in all_columns}
                full_row.update(overrides)
                return full_row

            if end > max_valid_date:
                rows.append(build_row({
                    "End Date": None,
                    "AbsenceDate_SAP": pd.NaT,
                    "Key_SAP": f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}",
                    "PY": "Python script - ORIGINAL"
                }))
            elif start == end:
                rows.append(build_row({
                    "Start Date": start,
                    "End Date": end,
                    "AbsenceDate_SAP": start,
                    "Key_SAP": f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}",
                    "PY": None
                }))
            else:
                rows.append(build_row({
                    "AbsenceDate_SAP": pd.NaT,
                    "Key_SAP": f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}",
                    "PY": "Python script - ORIGINAL"
                }))
                for d in pd.date_range(start, end):
                    rows.append(build_row({
                        "Start Date": d,
                        "End Date": d,
                        "AbsenceDate_SAP": d,
                        "Key_SAP": f"{row['Personnel Number']}_{d.strftime('%Y%m%d')}",
                        "PY": None
                    }))

        df = pd.DataFrame(rows)


        # === F5.2 - Mark which SAP rows are found in WD ===
        df["PY"] = df["Key_SAP"].isin(wd_df["Key_WD"]).map(
            {True: "OK", False: "Python script"}
        ).where(df["PY"] != "Python script - ORIGINAL", "Python script - ORIGINAL")


        # === F5.3 - Remove duplicates v2 (keep ORIGINAL at the bottom) ===
        df = df.sort_values(by="PY", ascending=True)
        df = df.drop_duplicates(subset=["Key_SAP"], keep="first")


        # === F5.4 - Replace ":  :" with "-" in 'Start' and 'End time' columns ===
        for col in ["Start", "End time"]:
            if col in df.columns:
                df[col] = df[col].astype(str).replace({":  :": "-"})

        # Remove duplicates v1
        #df = df.drop_duplicates(subset=["Key_SAP", "Status"], keep="first")
        #df = df.drop_duplicates(subset=["Key_SAP"], keep="first")
        
        
      # === F5.5 - Move columns to the end of the DataFrame ===
        cols = df.columns.tolist()
        for col in ["AbsenceDate_SAP", "Key_SAP", "PY"]:
            if col in cols:
                cols.remove(col)
        cols.extend(["AbsenceDate_SAP", "Key_SAP", "PY"])
        df = df[cols]


        # === F5.6 - Remove columns AbsenceDate_SAP/Key_SAP ===
        for col_to_remove in ["AbsenceDate_SAP", "Key_SAP"]:
            if col_to_remove in df.columns:
                df.drop(columns=[col_to_remove], inplace=True)


        # === F5.7 - Save to Excel ===
        log("📊 Saving results to Excel...")
        df.to_excel(out_path, index=False)


        # === F5.8 - Highlight rows with "ORIGINAL" status ===
        wb = load_workbook(out_path)
        ws = wb.active
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        PY_col = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}.get("PY", None)

        if PY_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[PY_col - 1].value == "Python script - ORIGINAL":
                    for cell in row:
                        cell.fill = fill
        wb.save(out_path)
        wb.close()

        # === F5.9 - Add formula columns (batch 1) after basic processing move batch 2 (AbsenceDate_SAP, Key_SAP, PY) to the end and remove the old columns ===
        add_formula_columns(out_path)
        reorder_columns(out_path)
            
        log("✅ Processing completed successfully.")
    except Exception as e:
        log(f"❌ ERROR during processing: {e}")


        
# === F6 - Wait until required files appear ===
def wait_for_files():
    log("🚀 Script started. Waiting for files")
    log("🔍 Watching for input files...")
    while True:
        files = os.listdir(watch_dir)
        if file_sap in files and file_wd in files:
            log("📂 Detected both files. Starting processing.")
            process_files()
            break
        time.sleep(check_interval)



# === F7 - Monitor folder and process go when files on place ===
if __name__ == "__main__":
    wait_for_files()
    