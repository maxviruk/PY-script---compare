import os
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

# === CONFIGURATION ===
watch_dir = os.path.join(os.getcwd(), "PY - Data")
file_sap = "Table_SAP.xlsx"
file_wd = "Table_WD.xlsx"
output_file = "AS01,AX04,AS03,AH01 - SAP_vs_WD.xlsx"
log_file = "processing_log.txt"
check_interval = 10

required_columns = [
    "Pers.No.", "Personnel Number", "EEGrp", "Employee Group", "S", "Employment Status",
    "CoCd", "Company Code", "PA", "Personnel Area", "ESgrp", "Employee Subgroup", "Start Date",
    "End Date", "Changed by", "Start", "End time", "A/AType", "Attendance or Absence Type"
]
max_valid_date = pd.Timestamp("2262-04-11")


def log(msg):
    with open(os.path.join(watch_dir, log_file), "a", encoding="utf-8") as logf:
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        logf.write(f"[{timestamp}] {msg}\n")
    print(f"[{timestamp}] {msg}")


def add_formula_columns(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    max_row = ws.max_row
    last_col_idx = ws.max_column

    new_columns = [
        ("Country ID",            '=LEFT(G{row};2)'),
        ("Country3",              '=VLOOKUP(G{row};Integration!$M:$O;3;0)'),
        ("SAP / non-SAP",         '=IF(AJ{row}="-";"-";VLOOKUP(AJ{row};Integration!$B:$I;8;0))'),
        ("#", '''=IF(AND(AK{row}="Non SAP Non Payroll systems";LEN(V{row})=0);"Non SAP Non Payroll systems";
IF(AK{row}="Non SAP Non Payroll systems";CONCAT(A{row};V{row};M{row});
IF(AK{row}="SAP Payroll systems";CONCAT(A{row};V{row};M{row});"-")))'''),
        ("##", '''=IF(AND(AK{row}="Non SAP Non Payroll systems";LEN(V{row})=0);"Non SAP Non Payroll systems";
IF(AK{row}="Non SAP Non Payroll systems";CONCAT(A{row};V{row};M{row};N{row});
IF(AK{row}="SAP Payroll systems";CONCAT(A{row};V{row};M{row};N{row});"-")))'''),
        ("CHECK", '''=IFERROR(
IF(YEAR(M{row})=2024;AL{row};
VLOOKUP(AL{row};WD!$BB:$BB;1;0));"-")'''),
        ("OK/NOK", '''=IFERROR(
IF(YEAR(M{row})=2024;"OK";
IF(OR(AN{row}<>"-";AP{row}="Integration");"OK";"NOK"));"NOK")'''),
        ("CHECK - Changed by - L1", None),
        ("CHECK - Changed by - L2", None),
        ("PY", None),
    ]

    for i, (col_name, formula_template) in enumerate(new_columns):
        col_idx = last_col_idx + 1 + i
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}1"] = col_name
        ws[f"{col_letter}1"].alignment = Alignment(horizontal='center')
        for row in range(2, max_row + 1):
            if col_name == "PY":
                ws[f"{col_letter}{row}"] = "Python script"
            elif formula_template:
                ws[f"{col_letter}{row}"] = formula_template.format(row=row)

    wb.save(file_path)
    wb.close()


def process_files():
    try:
        sap_path = os.path.join(watch_dir, file_sap)
        wd_path = os.path.join(watch_dir, file_wd)
        out_path = os.path.join(watch_dir, output_file)

        log("📂 Loading input files...")
        sap_df = pd.read_excel(sap_path)
        wd_df = pd.read_excel(wd_path)

        sap_df = sap_df[sap_df["A/AType"].isin(["AS01", "AX04", "AS03", "AH01"])].copy()
        all_columns = sap_df.columns.tolist()
        wd_df["Key_WD"] = wd_df["Employee ID"].astype(str) + "_" + wd_df["Time Off date"].dt.strftime("%Y%m%d")
        rows = []

        for _, row in sap_df.iterrows():
            start = row["Start Date"]
            end = row["End Date"]
            if pd.isnull(start) or pd.isnull(end):
                continue

            def build_row(overrides):
                full_row = {col: row[col] if col in required_columns else "-" for col in all_columns}
                full_row.update(overrides)
                return full_row

            if end > max_valid_date:
                rows.append(build_row({
                    "End Date": None,
                    "AbsenceDate_SAP": pd.NaT,
                    "Key_SAP": f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}",
                    "Status": "ORIGINAL"
                }))
            elif start == end:
                rows.append(build_row({
                    "Start Date": start,
                    "End Date": end,
                    "AbsenceDate_SAP": start,
                    "Key_SAP": f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}",
                    "Status": None
                }))
            else:
                rows.append(build_row({
                    "AbsenceDate_SAP": pd.NaT,
                    "Key_SAP": f"{row['Personnel Number']}_{start.strftime('%Y%m%d')}",
                    "Status": "ORIGINAL"
                }))
                for d in pd.date_range(start, end):
                    rows.append(build_row({
                        "Start Date": d,
                        "End Date": d,
                        "AbsenceDate_SAP": d,
                        "Key_SAP": f"{row['Personnel Number']}_{d.strftime('%Y%m%d')}",
                        "Status": None
                    }))

        df = pd.DataFrame(rows)
        df["Status"] = df["Key_SAP"].isin(wd_df["Key_WD"]).map({True: "OK", False: "Missing in WD"}).where(df["Status"] != "ORIGINAL", "ORIGINAL")
        df = df.sort_values(by="Status", ascending=True).drop_duplicates(subset=["Key_SAP"], keep="first")

        for col in ["Start", "End time"]:
            if col in df.columns:
                df[col] = df[col].astype(str).replace({":  :": "-"})

        if os.path.exists(out_path):
            date_suffix = datetime.now().strftime("%d%m")
            name, ext = os.path.splitext(output_file)
            output_file_with_date = f"{name}_{date_suffix}{ext}"
            out_path = os.path.join(watch_dir, output_file_with_date)
            log(f"📁 Output file exists. Saving as {output_file_with_date}")

        df.to_excel(out_path, index=False)

        wb = load_workbook(out_path)
        ws = wb.active
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        status_col = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}.get("Status", None)

        if status_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[status_col - 1].value == "ORIGINAL":
                    for cell in row:
                        cell.fill = fill
        wb.save(out_path)
        wb.close()

        # 👇 Добавляем формулы и колонки
        add_formula_columns(out_path)

        log("✅ Processing completed successfully.")

    except Exception as e:
        log(f"❌ ERROR during processing: {e}")


def wait_for_files():
    log("🔍 Watching for input files...")
    while True:
        files = os.listdir(watch_dir)
        if file_sap in files and file_wd in files:
            log("📂 Detected both files. Starting processing.")
            process_files()
            break
        time.sleep(check_interval)


if __name__ == "__main__":
    wait_for_files()
